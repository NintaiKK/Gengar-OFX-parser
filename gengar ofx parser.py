import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pdfplumber
from openpyxl import Workbook, load_workbook
import ofxparse
import os
from datetime import datetime

class ParserApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Conversor PDF/XLSX/OFX")
        self.root.geometry("650x450")
        
        self.pdf_path = tk.StringVar()
        self.xlsx_path = tk.StringVar()
        self.ofx_path = tk.StringVar()
        self.modo_captura = tk.StringVar(value="Padrão")  # Novo: Variável para o modo de captura
        
        self.create_widgets()
        
    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True)
        
        # Aba PDF para XLSX
        tab_pdf_xlsx = ttk.Frame(notebook)
        notebook.add(tab_pdf_xlsx, text="PDF para XLSX")
        
        ttk.Label(tab_pdf_xlsx, text="Arquivo PDF:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(tab_pdf_xlsx, textvariable=self.pdf_path, width=50).grid(row=0, column=1, pady=5)
        ttk.Button(tab_pdf_xlsx, text="Procurar", command=self.browse_pdf).grid(row=0, column=2, pady=5)
        
        ttk.Label(tab_pdf_xlsx, text="Arquivo XLSX de saída:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(tab_pdf_xlsx, textvariable=self.xlsx_path, width=50).grid(row=1, column=1, pady=5)
        ttk.Button(tab_pdf_xlsx, text="Procurar", command=self.browse_xlsx_output).grid(row=1, column=2, pady=5)
        
        ttk.Button(tab_pdf_xlsx, text="Converter PDF para XLSX", 
                  command=self.convert_pdf_to_xlsx).grid(row=2, column=1, pady=20)
        
        # Aba XLSX para OFX
        tab_xlsx_ofx = ttk.Frame(notebook)
        notebook.add(tab_xlsx_ofx, text="XLSX para OFX")
        
        # Widgets XLSX para OFX
        ttk.Label(tab_xlsx_ofx, text="Arquivo XLSX:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Entry(tab_xlsx_ofx, textvariable=self.xlsx_path, width=50).grid(row=0, column=1, pady=5)
        ttk.Button(tab_xlsx_ofx, text="Procurar", command=self.browse_xlsx).grid(row=0, column=2, pady=5)
        
        # Novo: Combobox para seleção do modo de captura
        ttk.Label(tab_xlsx_ofx, text="Modo de Captura:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.cb_modo_captura = ttk.Combobox(tab_xlsx_ofx, textvariable=self.modo_captura, 
                                          values=["Padrão", "Banco A", "Banco B", "Personalizado"])
        self.cb_modo_captura.grid(row=1, column=1, pady=5, sticky=tk.W)
        self.cb_modo_captura.current(0)
        
        ttk.Label(tab_xlsx_ofx, text="Arquivo OFX de saída:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(tab_xlsx_ofx, textvariable=self.ofx_path, width=50).grid(row=2, column=1, pady=5)
        ttk.Button(tab_xlsx_ofx, text="Procurar", command=self.browse_ofx_output).grid(row=2, column=2, pady=5)
        
        ttk.Button(tab_xlsx_ofx, text="Converter XLSX para OFX", 
                  command=self.convert_xlsx_to_ofx).grid(row=3, column=1, pady=20)
        
        self.status = ttk.Label(main_frame, text="Pronto", relief=tk.SUNKEN)
        self.status.pack(fill=tk.X, pady=10)
    
    def browse_pdf(self):
        filename = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
        if filename:
            self.pdf_path.set(filename)
            base = os.path.splitext(filename)[0]
            self.xlsx_path.set(base + ".xlsx")
    
    def browse_xlsx(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if filename:
            self.xlsx_path.set(filename)
            base = os.path.splitext(filename)[0]
            self.ofx_path.set(base + ".ofx")
    
    def browse_xlsx_output(self):
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                              filetypes=[("Excel Files", "*.xlsx")])
        if filename:
            self.xlsx_path.set(filename)
    
    def browse_ofx_output(self):
        filename = filedialog.asksaveasfilename(defaultextension=".ofx", 
                                              filetypes=[("OFX Files", "*.ofx")])
        if filename:
            self.ofx_path.set(filename)
    
    def convert_pdf_to_xlsx(self):
        pdf_file = self.pdf_path.get()
        xlsx_file = self.xlsx_path.get()
        
        if not pdf_file or not xlsx_file:
            messagebox.showerror("Erro", "Por favor, especifique os arquivos de entrada e saída")
            return
        
        try:
            self.status.config(text="Convertendo PDF para XLSX...")
            self.root.update()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Dados do PDF"
            
            with pdfplumber.open(pdf_file) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        for line in text.split('\n'):
                            ws.append([line])
            
            wb.save(xlsx_file)
            
            messagebox.showinfo("Sucesso", f"Arquivo convertido com sucesso para:\n{xlsx_file}")
            self.status.config(text="Pronto")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro durante a conversão:\n{str(e)}")
            self.status.config(text="Erro")
    
    def convert_xlsx_to_ofx(self):
        xlsx_file = self.xlsx_path.get()
        ofx_file = self.ofx_path.get()
        
        if not xlsx_file or not ofx_file:
            messagebox.showerror("Erro", "Por favor, especifique os arquivos de entrada e saída")
            return
        
        try:
            self.status.config(text="Convertendo XLSX para OFX...")
            self.root.update()
            
            # Carregar o arquivo Excel
            wb = load_workbook(xlsx_file, read_only=True)
            ws = wb.active
            
            # Obter a data atual no formato OFX
            current_date = datetime.now().strftime("%Y%m%d")
            
            # Criar cabeçalho OFX
            ofx_content = f"""OFXHEADER:100
DATA:OFXSGML
VERSION:102
SECURITY:NONE
ENCODING:USASCII
CHARSET:1252
COMPRESSION:NONE
OLDFILEUID:NONE
NEWFILEUID:NONE

<OFX>
    <SIGNONMSGSRSV1>
        <SONRS>
            <STATUS>
                <CODE>0
                <SEVERITY>INFO
            </STATUS>
            <DTSERVER>{current_date}
            <LANGUAGE>POR
        </SONRS>
    </SIGNONMSGSRSV1>
    <BANKMSGSRSV1>
        <STMTTRNRS>
            <TRNUID>1
            <STATUS>
                <CODE>0
                <SEVERITY>INFO
            </STATUS>
            <STMTRS>
                <CURDEF>BRL
                <BANKACCTFROM>
                    <BANKID>123
                    <ACCTID>456789
                    <ACCTTYPE>CHECKING
                </BANKACCTFROM>
                <BANKTRANLIST>
                    <DTSTART>{current_date}
                    <DTEND>{current_date}
"""
            
            # Processar linhas do Excel de acordo com o modo selecionado
            transactions = []
            modo = self.modo_captura.get()
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if not row or not any(row):
                    continue
                
                try:
                    # Diferentes modos de captura
                    if modo == "Padrão":
                        # Modo padrão: colunas Data, Descrição, Valor
                        trans_date = row[0].strftime("%Y%m%d") if hasattr(row[0], 'strftime') else current_date
                        description = str(row[1]) if len(row) > 1 else "Transação"
                        amount = float(row[2]) if len(row) > 2 and row[2] is not None else 0.0
                    
                    elif modo == "Banco A":
                        # Modo para Banco A: colunas Valor, Data, Histórico
                        amount = float(row[0]) if row[0] is not None else 0.0
                        trans_date = row[1].strftime("%Y%m%d") if hasattr(row[1], 'strftime') else current_date
                        description = str(row[2]) if len(row) > 2 else "Transação"
                    
                    elif modo == "Banco B":
                        # Modo para Banco B: colunas Data, Valor, Tipo, Descrição
                        trans_date = row[0].strftime("%Y%m%d") if hasattr(row[0], 'strftime') else current_date
                        amount = float(row[1]) if row[1] is not None else 0.0
                        trans_type = "CREDIT" if str(row[2]).upper() in ["C", "CREDITO", "CRÉDITO"] else "DEBIT"
                        description = str(row[3]) if len(row) > 3 else "Transação"
                    
                    elif modo == "Personalizado":
                        # Modo personalizado: captura todas as colunas como MEMO
                        trans_date = current_date
                        amount = 0.0
                        description = " | ".join([str(cell) if cell is not None else "" for cell in row])
                    
                    transaction = f"""                    <STMTTRN>
                        <TRNTYPE>{"DEBIT" if amount < 0 else "CREDIT"}
                        <DTPOSTED>{trans_date}
                        <TRNAMT>{amount:.2f}
                        <FITID>{row_idx}
                        <MEMO>{description}
                    </STMTTRN>
"""
                    transactions.append(transaction)
                except (ValueError, IndexError, AttributeError) as e:
                    print(f"Erro ao processar linha {row_idx} no modo {modo}: {e}")
                    continue
            
            # Juntar todas as transações
            ofx_content += "".join(transactions)
            
            # Rodapé OFX
            ofx_content += f"""                </BANKTRANLIST>
                <LEDGERBAL>
                    <BALAMT>0.00
                    <DTASOF>{current_date}
                </LEDGERBAL>
            </STMTRS>
        </STMTTRNRS>
    </BANKMSGSRSV1>
</OFX>"""
            
            # Escrever o arquivo OFX
            with open(ofx_file, 'w', encoding='utf-8') as f:
                f.write(ofx_content)
            
            messagebox.showinfo("Sucesso", f"Arquivo convertido com sucesso para:\n{ofx_file}")
            self.status.config(text="Pronto")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro durante a conversão:\n{str(e)}")
            self.status.config(text="Erro")

if __name__ == "__main__":
    root = tk.Tk()
    app = ParserApp(root)
    root.mainloop()
